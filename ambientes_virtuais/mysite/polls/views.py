import logging,os
from django.shortcuts import render, redirect
from .forms import ClienteForm, LoginForm
from django.db import IntegrityError
from django.contrib import messages
import pandas as pd
from .models import Cliente,framework, ActionModel
from django.contrib.auth import logout as django_logout
from django.utils import timezone
from django.contrib.auth import authenticate, login as auth_login
import openpyxl
from django.http import HttpResponse
from datetime import datetime
from .models import TemporaryActionModel,Log
import plotly.graph_objs as go
import plotly.graph_objects as go
from io import BytesIO
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.contrib.auth import login as auth_login
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.models import User
from django.utils.http import urlsafe_base64_decode
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.views import PasswordResetView
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from openpyxl import Workbook

logger = logging.getLogger('django')
logger = logging.getLogger(__name__)


@login_required(login_url='login')
def index(request):
    column_names = [
        'Control', 'Sub-Control', 'Ativo', 'Função de segurança', 'Título', 'Descrição',
        'NIST CSF', 'IG', 'Nome da subcategoria'
    ]

    # Inicialização das variáveis
    grafico_velocimetro = None
    grafico_controle = None
    ig_data = {}  # Para conter as porcentagens e metas
    asset_type_counts = {}
    datas_uploads = []
    selected_date = None

    # Recupera o cliente pelo usuário autenticado
    try:
        user = request.user
        # Obtém o nome do cliente do usuário autenticado
        first_name = user.first_name
        last_name = user.last_name
        nome_cliente = f'{first_name} {last_name}' if first_name else 'Desconhecido'

       
        datas_uploads = data_consulta_grafico(request)

        if request.method == 'POST':
            selected_date = request.POST.get('selected_date')  # Data selecionada no formulário

            if selected_date:
                # Filtra as ações pelo cliente e data selecionada
                actions = ActionModel.objects.filter(upload_date=selected_date, nome=nome_cliente)

                # Gráfico de velocímetro
                sim_count = actions.filter(acao='sim').count()
                nao_count = actions.filter(acao='nao').count()
                grafico_velocimetro = cria_grafico_velocimetro(sim_count, nao_count)

                # Calcula porcentagens e metas para cada IG
                ig_data = calcula_porcentagem_ig(actions)

                # Calcula contagem de tipos de ativo
                asset_type_counts = calcula_ativos_grafico(actions)

                # Gráfico de linha por controle
                grafico_controle = cria_grafico_controle(actions)

                # Adiciona o log informando que o cliente gerou os gráficos
                Log.objects.create(
                    cliente=user,
                    descricao=f'O cliente {user.first_name} {user.last_name} gerou gráficos para a data {selected_date}.'
                )

    except Cliente.DoesNotExist:
        # Caso o cliente não exista, as variáveis permanecerão vazias
        pass

    # Dados para a tabela principal
    data_with_actions = [
        {
            'id': item.id,
            'cis_control': item.cis_control,
            'cis_sub_control': item.cis_sub_control,
            'tipo_de_ativo': item.tipo_de_ativo,
            'funcao_de_seguranca': item.funcao_de_seguranca,
            'titulo': item.titulo,
            'descricao': item.descricao,
            'nist_csf': item.nist_csf,
            'ig': item.ig,
            'nome_da_subcategoria': item.nome_da_subcategoria,
            'acao': ''
        }
        for item in framework.objects.all()
    ]

    
    return render(request, 'index.html', {
        'data': data_with_actions,
        'column_names': column_names,
        'datas_uploads': datas_uploads,
        'selected_date': selected_date,
        'grafico_velocimetro': grafico_velocimetro,
        'grafico_controle': grafico_controle,
        'ig_data': ig_data,
        'asset_type_counts': asset_type_counts,
    })





def logout_view(request):
    # Obter o cliente associado ao logout
    cliente_id = request.session.get('cliente_id')


    if cliente_id:
        try:
            cliente = Cliente.objects.get(id=cliente_id)
            user = cliente.user  # Obter o usuário associado ao cliente

            if user:
                # Registrar o logout no modelo Log
                Log.objects.create(
                    cliente=user,
                    descricao=f'O cliente {user.first_name} {user.last_name} realizou logout.'
                )
                
            else:
                logger.warning(f'O cliente com ID {cliente_id} não possui um usuário associado.')
        except Cliente.DoesNotExist:
            logger.error(f'Cliente com ID {cliente_id} não encontrado.')
    else:
        logger.warning(f"cliente_id não encontrado na sessão. O cliente pode não estar logado.")

    # Encerrar a sessão do usuário
    django_logout(request)

    # Remover o cliente_id da sessão
    request.session.pop('cliente_id', None)

    return redirect('login')


def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            # Tenta autenticar o usuário com o email e senha
            user = authenticate(request, username=email, password=password)

            if user is not None:
                # Realiza o login
                auth_login(request, user)

                # Armazena o cliente associado ao usuário na sessão
                try:
                    cliente = Cliente.objects.get(user=user)
                    request.session['cliente_id'] = cliente.id

                    # Registra o login do cliente na tabela de logs
                    Log.objects.create(
                        cliente=user,
                        descricao=f'O cliente {user.first_name} {user.last_name} realizou login.'
                    )


                except Cliente.DoesNotExist:
                    messages.error(request, "Erro: Cliente associado ao usuário não encontrado.")
                    logger.error(f"Cliente associado ao usuário {user.username} não encontrado.")
                    return redirect('login')

                return redirect('index')  # Página inicial após o login
            else:
                messages.error(request, 'Email ou senha incorretos.')
                logger.warning(f"Tentativa de login falhou para o email: {email}")
        else:
            messages.error(request, 'Formulário inválido. Verifique os dados informados.')
            logger.warning("Formulário de login inválido enviado.")
    else:
        form = LoginForm()

    return render(request, 'login.html', {'form': form})


def registro_view(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            try:
                cliente = form.save()  # Salva o cliente no banco de dados
                
                # Autentica o usuário utilizando o backend padrão
                user = cliente.user
                backend = 'django.contrib.auth.backends.ModelBackend'  # Caminho do backend como string
                auth_login(request, user, backend=backend)  # Passa o backend explicitamente
                
                # Armazena o cliente na sessão
                request.session['cliente_id'] = cliente.id

                # Registra o evento de criação na tabela de logs
                Log.objects.create(
                    cliente=user,
                    descricao=f'O cliente {user.first_name} {user.last_name} realizou registro.'
                )


                return redirect('index')  # Redireciona para a página inicial
            except IntegrityError:
                form.add_error(None, "Erro: O email ou CNPJ já está registrado.")
                logger.warning("Tentativa de registro falhou: email ou CNPJ já está registrado.")
    else:
        form = ClienteForm()

    return render(request, 'register.html', {'form': form})


def custom_password_reset_confirm(request, uidb64, token):
    try:
        # Decodificando o ID do usuário a partir do token
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    # Se o método for POST, processa a alteração de senha
    if request.method == "POST":
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        # Verificando se as senhas coincidem
        if password != confirm_password:
            messages.error(request, "As senhas não coincidem.")
            # Log de erro ao tentar redefinir senha com senhas não coincidentes
            if user:
                Log.objects.create(
                    cliente=user,
                    descricao=f"Tentativa de redefinir senha com senhas não coincidentes para {user.first_name} {user.last_name}."
                )
            return render(request, 'registration/solicitando_nova_senha.html', {'uid': uidb64, 'token': token})

        # Verificando se o token é válido
        if user and default_token_generator.check_token(user, token):
            user.set_password(password)
            user.save()
            update_session_auth_hash(request, user)  # Mantém o usuário autenticado após a troca de senha
            # Log de sucesso ao alterar a senha
            Log.objects.create(
                cliente=user,
                descricao=f"A senha foi alterada com sucesso para {user.first_name} {user.last_name}."
            )
            return redirect('confirmacao_senha_alterada')
        else:
            messages.error(request, "O link para redefinir a senha é inválido ou expirou.")
            # Log de erro quando o token for inválido
            if user:
                Log.objects.create(
                    cliente=user,
                    descricao=f"Tentativa de redefinir senha falhou para {user.first_name} {user.last_name} devido a um token inválido ou expirado."
                )

    # Retorna o formulário de redefinição de senha se não for POST ou se houver erro
    return render(request, 'registration/solicitando_nova_senha.html', {'uid': uidb64, 'token': token})


class CustomPasswordResetView(PasswordResetView):
    def form_valid(self, form):
        email = form.cleaned_data['email']

        # Validação do formato do e-mail
        try:
            validate_email(email)
        except ValidationError:
            form.add_error('email', "E-mail inválido. Verifique e tente novamente.")
            return self.form_invalid(form)

        # Verifica se o e-mail existe exatamente no banco de dados
        if not User.objects.filter(email=email).exists():
            form.add_error('email', "O e-mail fornecido não está cadastrado no sistema.")
            return self.form_invalid(form)

        # Marca a sessão como válida para acesso às páginas subsequentes
        self.request.session['password_reset_allowed'] = True
        # Log de sucesso ao tentar redefinir senha
        user = User.objects.get(email=email)
        Log.objects.create(
            cliente=user,
            descricao=f"Solicitação de redefinição de senha para {user.first_name} {user.last_name} com e-mail {email}."
        )

        return super().form_valid(form)





def data_consulta_grafico(request):
    """Obtém datas de upload únicas associadas ao cliente no formato d/m/Y para exibição."""
    
    # Obtém o nome completo do cliente do usuário autenticado
    user = request.user
    first_name = 'Desconhecido'
    last_name = ''
    
    if user.is_authenticated:
        first_name = user.first_name
        last_name = user.last_name
    
    # Nome do cliente no formato 'Primeiro Nome Sobrenome'
    nome_cliente = f'{first_name} {last_name}' if first_name else 'Desconhecido'
    
    # Filtra as datas de upload únicas para o nome do cliente
    return [
        {
            'value': date.strftime('%Y-%m-%d'),       # Formato original para envio
            'display': date.strftime('%d/%m/%Y')      # Formato d/m/Y para exibição
        }
        for date in ActionModel.objects.filter(nome=nome_cliente)
                                       .values_list('upload_date', flat=True)
                                       .distinct().order_by('upload_date')
    ]


def cria_grafico_velocimetro(sim_count, nao_count):
    """Cria o gráfico de velocímetro baseado na porcentagem de ações 'Sim'."""
    total_actions = sim_count + nao_count
    percentage = (sim_count / total_actions) * 100 if total_actions > 0 else 0
    fig_velocimetro = go.Figure(go.Indicator(
        mode="gauge+number", value=percentage,
        title={'text': "Velocímetro", 'font': {'size': 16}},
        number={'font': {'size': 28, 'color': 'blue'}},
        gauge={
            'axis': {'range': [0, 100], 'tickwidth': 1, 'tickcolor': "black"},
            'bar': {'color': "darkblue", 'thickness': 0.2},
            'bgcolor': "white", 'borderwidth': 1, 'bordercolor': "gray",
            'steps': [
                {'range': [0, 50], 'color': "LightBlue"},
                {'range': [50, 100], 'color': "RoyalBlue"}
            ],
            'threshold': {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': 90}
        }
    ))
    fig_velocimetro.update_layout(
        title_text="Percentual de Ações 'Sim'", titlefont_size=16,
        margin=dict(l=35, r=35, t=30, b=10), font=dict(color='black', size=12)
    )
    return fig_velocimetro.to_html(full_html=False)


def calcula_porcentagem_ig(actions):
    ig_data = {}

    # Obtém os valores distintos de IG, excluindo os valores None, e ordena
    ig_values = actions.values_list('ig', flat=True).exclude(ig=None).distinct().order_by('ig')

    for ig in ig_values:
        # Total de ações para o IG atual (inclui "Não se aplica")
        ig_total = actions.filter(ig=ig).count()

        # Contagem de "Sim" para o IG atual
        ig_sim_count = actions.filter(ig=ig, acao='sim').count()

        # Calcula a porcentagem de "Sim" para o IG atual
        ig_percentage = (ig_sim_count / ig_total) * 100 if ig_total > 0 else 0

        # Total de ações válidas para a meta (exclui "Não se aplica")
        ig_total_valid = actions.filter(ig=ig).exclude(acao='nao se aplica').count()

        # Calcula a meta para o IG atual
        ig_meta = (ig_total_valid / ig_total) * 100 if ig_total > 0 else 0

        ig_data[ig] = {
            "percentage": round(ig_percentage, 2),
            "meta": round(ig_meta, 2),
        }

    return ig_data


def calcula_ativos_grafico(actions):
    """Calcula a contagem de 'Sim' e o total para cada tipo de ativo."""
    asset_type_counts = {}

    # Obtém os tipos de ativo distintos
    asset_types = actions.values_list('tipo_de_ativo', flat=True).distinct()

    for tipo in asset_types:
        # Conta quantos registros têm 'Sim' e o total, ignorando 'Não se aplica'
        tipo_sim_count = actions.filter(tipo_de_ativo=tipo, acao='sim').count()
        tipo_total = actions.filter(tipo_de_ativo=tipo).exclude(acao='nao se aplica').count()
        
        asset_type_counts[tipo] = {
            'sim_count': tipo_sim_count,
            'total': tipo_total
        }

    return asset_type_counts


def cria_grafico_controle(actions):
    controls = list(range(1, 21))
    control_titles = [
        "   Inventário e Controle de<br>   Ativos de Hardware", "   Inventário e Controle de<br>   Ativos de Software",
        "   Gerenciamento contínuo<br>   de vulnerabilidades", "   Uso controlado de<br>   privilégios administrativos",
        "   Configuração segura para hardware<br>   e software em dispositivos móveis,<br>   laptops,estações de trabalho e servidores",
        "   Manutenção, Monitoramento e Análise de <br>   registros de Auditoria", "   Proteção de e-mail e<br>   navegador da Web",
        "   Defesas contra malware", "   Limitação e Controle de Portas,<br>   Protocolos e<br>   Serviços de Rede",
        "   Capacidades de<br>   recuperação de dados", "   Configuração Segura para Rede Dispositivos,<br>   como Firewalls, Roteadores e Switches",
        "   Defesa de Limites", "   Proteção de Dados", "   Acesso controlado com<br>   base na necessidade de saber",
        "   Controle de acesso<br>   sem fio", "   Monitoramento e Controle<br>   de Contas",
        "   Implementar um programa de<br>   treinamento e conscientização<br>   de segurança", "   Segurança de Software<br>   de Aplicação",
        "   Resposta e Gerenciamento<br>   de Incidentes", "   Testes de Penetração<br>   e Exercícios de Equipe Vermelha"
    ]

    # Calcula a meta para cada controle como (total de campos - campos "não se aplica")
    metas = []
    sim_counts = []
    for control in controls:
        total_campos = actions.filter(cis_control=control).count()
        nao_se_aplica = actions.filter(cis_control=control, acao='não se aplica').count()
        meta = total_campos - nao_se_aplica
        metas.append(meta)
        
        # Conta o número de "Sim" para cada controle
        sim_count = actions.filter(cis_control=control, acao='sim').count()
        sim_counts.append(sim_count)

    # Cria o gráfico de barras
    fig = go.Figure()

    # Adiciona as barras com os valores "Sim"
    fig.add_trace(go.Bar(
        x=control_titles,
        y=sim_counts,
        text=sim_counts,
        textposition='inside',
        insidetextanchor='start',  
        name='Aderência',
        marker=dict(color="rgb(0, 51, 102)", line=dict(color="rgb(0, 51, 102)", width=2)),
        textfont=dict(size=14, color="white")  
    ))

    # Adiciona a linha de meta para cada controle
    fig.add_trace(go.Scatter(
        x=control_titles,
        y=metas,
        mode='lines',
        name='Meta',
        line=dict(color='rgb(78, 177, 210)', width=2)  
    ))

    # Adiciona os valores das metas acima da linha, com um pequeno deslocamento ajustando o valor de y
    fig.add_trace(go.Scatter(
        x=control_titles,
        y=[meta + 0.5 for meta in metas],  # Adiciona 2 de deslocamento aos valores das metas
        mode='text',
        text=metas,
        textposition='top center',
        showlegend=False,  
        textfont=dict(size=12, color="black")
    ))

    # Configurações do layout
    fig.update_layout(
        title={
            'text': "Aderência do CIS Controls por Categoria vs Meta",
            'x': 0.5,  # Centraliza o título
            'xanchor': 'center',
            'font': {'size': 20}  
        },
        #xaxis_title="Controle",
        #yaxis_title="Número de 'Sim'",
        yaxis=dict(range=[0, max(max(sim_counts), max(metas)) + 5]),  # Ajusta o limite do eixo y
        showlegend=True,
        barmode='group',
        bargap=0.5,  # Aumenta o espaçamento entre as barras
        template="plotly_white",
        width=1100,  # Aumenta a largura do gráfico
        height=800   # Aumenta a altura do gráfico
    )

    # Ajusta a rotação, o alinhamento e o espaçamento dos títulos no eixo X
    fig.update_xaxes(
        tickangle=90,  
        tickfont=dict(size=14),  
        tickvals=control_titles,
        ticktext=control_titles,
        titlefont=dict(size=16),  
        title_standoff=50,  
        
    )

    # Ajusta as margens para garantir que os rótulos não fiquem grudados
    fig.update_layout(margin=dict(l=50, r=40, t=80, b=150))  # Aumenta a margem inferior

    return fig.to_html(full_html=False)





def upload_excel(request):
    # Verificar se o cliente_id está na sessão
    cliente_id = request.session.get('cliente_id')

    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'Nenhum arquivo selecionado.')
            return redirect('index')
        
        excel_file = request.FILES['file']
        if not excel_file:
            messages.error(request, 'Nenhum arquivo selecionado.')
            return redirect('index')

        # Caminho para a pasta 'FrameWork' dentro da pasta 'media'
        framework_directory = os.path.join(settings.MEDIA_ROOT, 'FrameWork')

        # Verificar se a pasta 'FrameWork' existe, caso contrário, criá-la
        if not os.path.exists(framework_directory):
            os.makedirs(framework_directory)

        # Remover qualquer arquivo existente dentro da pasta 'FrameWork'
        for f in os.listdir(framework_directory):
            file_path = os.path.join(framework_directory, f)
            if os.path.isfile(file_path):
                os.remove(file_path)

        # Definir o caminho final para salvar o arquivo
        filename = os.path.join(framework_directory, excel_file.name)

        # Salvar o arquivo Excel na pasta 'FrameWork'
        with open(filename, 'wb') as f:
            for chunk in excel_file.chunks():
                f.write(chunk)

        # Ler o arquivo Excel
        try:
            df = pd.read_excel(filename)

            # Substituir NaN por strings vazias
            df.fillna('', inplace=True)

            # Excluir dados antigos no banco de dados (se necessário)
            framework.objects.all().delete()

            # Mapear as colunas do DataFrame para campos do modelo
            column_mapping = {
                'CIS Control': 'cis_control',
                'CIS Sub-Control': 'cis_sub_control',
                'Tipo de ativo': 'tipo_de_ativo',
                'Função de segurança': 'funcao_de_seguranca',
                'Título': 'titulo',
                'Descrição': 'descricao',
                'NIST CSF': 'nist_csf',
                'IG': 'ig',  # Adicione o mapeamento da coluna IG
                'Nome da subcategoria': 'nome_da_subcategoria',
            }

            # Verificar se as colunas esperadas estão presentes no DataFrame
            for col in column_mapping.keys():
                if col not in df.columns:
                    messages.error(request, f'Coluna {col} não encontrada no arquivo Excel.')
                    return redirect('index')

            # Salvar os dados no banco de dados
            for index, row in df.iterrows():
                framework.objects.create(
                    cis_control=row['CIS Control'],
                    cis_sub_control=row['CIS Sub-Control'],
                    tipo_de_ativo=row['Tipo de ativo'],
                    funcao_de_seguranca=row['Função de segurança'],
                    titulo=row['Título'],
                    descricao=row['Descrição'],
                    nist_csf=row['NIST CSF'],
                    ig=row['IG'], # Adicione IG aqui
                    nome_da_subcategoria=row['Nome da subcategoria'],
                )

            # Mensagem de sucesso
            messages.success(request, 'Arquivo enviado com sucesso!')

            # Registrar o evento de upload no modelo Log
            cliente_id = request.session.get('cliente_id')
            if cliente_id:
                try:
                    cliente = Cliente.objects.get(id=cliente_id)
                    user = cliente.user  # Obter o usuário associado ao cliente
                    Log.objects.create(
                        cliente=user,
                        descricao=f'O cliente {user.first_name} {user.last_name} fez upload de um FrameWork.'
                    )
                except Cliente.DoesNotExist:
                    messages.error(request, "Erro: Cliente associado ao usuário não encontrado.")
                    return redirect('index')
            else:
                messages.error(request, "Erro: Cliente não encontrado na sessão.")
                return redirect('index')

        except Exception as e:
            messages.error(request, f'Ocorreu um erro ao processar o arquivo: {str(e)}')
            return redirect('index')

        return redirect('index')
    
    return redirect('index')


def atualiza_tabela(request):
    if request.method == 'POST':
        try:
            today = timezone.now().date()
            user = request.user
            first_name = 'Desconhecido'
            last_name = ''

            if user.is_authenticated:
                first_name = user.first_name
                last_name = user.last_name

            # Verifica se o cliente existe
            nome_cliente = f'{first_name} {last_name}' if first_name else 'Desconhecido'

            # Logar a atualização de linha
            if request.POST.get('save') == 'temporary':
                item_id = request.POST.get('editId')
                new_action = request.POST.get(f'action_{item_id}')
                
                if item_id and new_action:
                    try:
                        framework_item = framework.objects.get(id=item_id)
                    except framework.DoesNotExist:
                        messages.error(request, 'Item do framework não encontrado.')
                        return redirect('index')

                    TemporaryActionModel.objects.update_or_create(
                        nome=nome_cliente,
                        titulo=framework_item.titulo,
                        defaults={
                            'acao': new_action,
                            'cis_control': framework_item.cis_control,
                            'cis_sub_control': framework_item.cis_sub_control,
                            'tipo_de_ativo': framework_item.tipo_de_ativo,
                            'funcao_de_seguranca': framework_item.funcao_de_seguranca,
                            'descricao': framework_item.descricao,
                            'nist_csf': framework_item.nist_csf,
                            'ig': framework_item.ig,
                            'nome_da_subcategoria': framework_item.nome_da_subcategoria,
                            'upload_date': today
                        }
                    )
                    messages.success(request, 'Ação da linha salva temporariamente.')

                    # Log para atualização de linha
                    Log.objects.create(
                        cliente=user,
                        descricao=f'O cliente {user.first_name} {user.last_name} atualizou a linha da sua tabela {framework_item.titulo}.'
                    )

                else:
                    messages.error(request, 'Nenhuma linha foi selecionada para atualizar.')

            elif request.POST.get('save') == 'final':
                data_for_excel = []
                saved_count = 0

                for item in framework.objects.all():
                    action = request.POST.get(f'action_{item.id}', '')
                    action_obj, created = ActionModel.objects.update_or_create(
                        upload_date=today, nome=nome_cliente, titulo=item.titulo, cis_sub_control=item.cis_sub_control,
                        defaults={
                            'cis_control': item.cis_control,
                            'tipo_de_ativo': item.tipo_de_ativo,
                            'funcao_de_seguranca': item.funcao_de_seguranca,
                            'descricao': item.descricao,
                            'nist_csf': item.nist_csf,
                            'ig': item.ig,
                            'nome_da_subcategoria': item.nome_da_subcategoria,
                            'acao': action
                        }
                    )
                    if created:
                        saved_count += 1

                    data_for_excel.append({
                        'cis_control': item.cis_control,
                        'cis_sub_control': item.cis_sub_control,
                        'tipo_de_ativo': item.tipo_de_ativo,
                        'funcao_de_seguranca': item.funcao_de_seguranca,
                        'titulo': item.titulo,
                        'descricao': item.descricao,
                        'nist_csf': item.nist_csf,
                        'ig': item.ig,
                        'nome_da_subcategoria': item.nome_da_subcategoria,
                        'acao': action
                    })

                # Gerar e salvar o arquivo Excel
                df = pd.DataFrame(data_for_excel)
                df.fillna('', inplace=True)

                filename = f"{nome_cliente}_{today}.xlsx"
                file_path = f"{filename}"

                if default_storage.exists(file_path):
                    default_storage.delete(file_path)

                excel_file = BytesIO()
                df.to_excel(excel_file, index=False)
                excel_file.seek(0)
                default_storage.save(file_path, ContentFile(excel_file.read()))

                messages.success(request, 'Tabela enviada com sucesso!')

                # Log para envio da tabela
                Log.objects.create(
                    cliente=user,
                    descricao=f'O cliente {user.first_name} {user.last_name} enviou a sua tabela.'
                )

            else:
                TemporaryActionModel.objects.filter(nome=nome_cliente).delete()
                
                for item in framework.objects.all():
                    action = request.POST.get(f'action_{item.id}', '')
                    TemporaryActionModel.objects.create(
                        nome=nome_cliente,
                        cis_control=item.cis_control,
                        cis_sub_control=item.cis_sub_control,
                        tipo_de_ativo=item.tipo_de_ativo,
                        funcao_de_seguranca=item.funcao_de_seguranca,
                        titulo=item.titulo,
                        descricao=item.descricao,
                        nist_csf=item.nist_csf,
                        ig=item.ig,
                        nome_da_subcategoria=item.nome_da_subcategoria,
                        acao=action,
                        upload_date=today
                    )

                messages.success(request, 'Tabela salva temporariamente com sucesso!')

                # Log para salvar tabela temporariamente
                Log.objects.create(
                    cliente=user,
                    descricao=f'O cliente {user.first_name} {user.last_name} salvou a sua tabela.'
                )

        except Exception as e:
            messages.error(request, f'Ocorreu um erro ao atualizar a tabela: {str(e)}')

        return redirect('index')
    else:
        return redirect('index')


def baixar_tabela(request):
    # Recebe parâmetros da requisição
    user_name = request.GET.get('user_name')
    submission_date_str = request.GET.get('submission_date')

    # Verifica se os parâmetros foram fornecidos
    if not user_name or not submission_date_str:
        logger.error(f"Nome do usuário ou data não fornecidos. user_name: {user_name}, submission_date: {submission_date_str}")
        messages.error(request, "Nome do usuário ou data não fornecidos.")
        return redirect('index')

    # Converte a string de data do formato YYYY-MM-DD para um objeto de data
    try:
        submission_date = datetime.strptime(submission_date_str, '%Y-%m-%d').date()
    except ValueError:
        logger.error(f"Data no formato inválido fornecida. Data recebida: {submission_date_str}")
        messages.error(request, "Data no formato inválido. Use o formato YYYY-MM-DD.")
        return redirect('index')

    # Busca os registros que correspondem ao nome e data fornecidos
    registros = ActionModel.objects.filter(
        nome=user_name,
        upload_date=submission_date  # Comparando a data completa, sem hora
    )

    if not registros.exists():
        logger.warning(f"Nenhum registro encontrado para o usuário {user_name} na data {submission_date}.")
        messages.error(request, f"Nenhum registro encontrado para o usuário {user_name} na data {submission_date}.")
        return redirect('index')

    # Cria um novo workbook e sheet
    wb = Workbook()
    ws = wb.active

    # Garantir que o nome da planilha não tenha mais de 31 caracteres
    sheet_title = f"Planilha de {user_name}"
    ws.title = sheet_title[:31]  # Truncar o título para 31 caracteres

    # Especifica os cabeçalhos das colunas
    headers = [
        "CIS Control", 
        "CIS Sub-Control", 
        "Tipo de ativo", 
        "Função de segurança", 
        "Título", 
        "Descrição",
        "NIST CSF",
        "IG",
        "Nome da subcategoria",
        "Ação"
    ]

    # Insere os cabeçalhos na primeira linha
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Preenche a planilha com os dados do banco de dados
    for row_num, registro in enumerate(registros, 2):
        ws.cell(row=row_num, column=1, value=registro.cis_control)
        ws.cell(row=row_num, column=2, value=registro.cis_sub_control)
        ws.cell(row=row_num, column=3, value=registro.tipo_de_ativo)
        ws.cell(row=row_num, column=4, value=registro.funcao_de_seguranca)
        ws.cell(row=row_num, column=5, value=registro.titulo)
        ws.cell(row=row_num, column=6, value=registro.descricao)
        ws.cell(row=row_num, column=7, value=registro.nist_csf)
        ws.cell(row=row_num, column=8, value=registro.ig)
        ws.cell(row=row_num, column=9, value=registro.nome_da_subcategoria)
        ws.cell(row=row_num, column=10, value=registro.acao)

    # Criar a resposta HTTP com o arquivo gerado
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={user_name}_{submission_date}.xlsx'
    nome_tabela = f"{user_name}_{submission_date}.xlsx"
    # Tente forçar a gravação no response
    try:
        wb.save(response)
        # Log para o sucesso do download da tabela
        Log.objects.create(
            cliente=request.user,
            descricao=f'O cliente {request.user.first_name} {request.user.last_name} baixou a tabela {nome_tabela}.'
        )
        
    except Exception as e:
        logger.error(f"Erro ao salvar o arquivo Excel: {str(e)}")
        messages.error(request, "Erro ao gerar o arquivo Excel.")
        return redirect('index')

    return response


def carrega_tabela_temporaria(request):
    # Verifica se o usuário está autenticado
    if request.user.is_authenticated:
        try:
            user = request.user
            first_name = 'Desconhecido'
            last_name = ''

            # Obtém o nome completo do cliente
            if user.is_authenticated:
                first_name = user.first_name
                last_name = user.last_name

            # Nome do cliente no formato 'Primeiro Nome Sobrenome'
            nome_cliente = f'{first_name} {last_name}' if first_name else 'Desconhecido'

            temp_records = TemporaryActionModel.objects.filter(nome=nome_cliente)

            # Verifica se existem registros temporários
            if not temp_records.exists():
                messages.info(request, 'Nenhum registro temporário encontrado.')
                return redirect('index')

            # Carrega os dados do framework
            data = framework.objects.all()

            # Define os nomes das colunas manualmente
            column_names = [
                'Control', 'Sub-Control', 'Ativo',
                'Função de segurança', 'Título', 'Descrição',
                'NIST CSF','IG','Nome da subcategoria'
            ]

            # Mapeia as ações temporárias por título
            temp_actions = {record.titulo: record.acao for record in temp_records}

            # Cria uma nova lista de dados com o campo 'acao'
            data_with_actions = []
            for item in data:
                data_with_actions.append({
                    'id': item.id,
                    'cis_control': item.cis_control,
                    'cis_sub_control': item.cis_sub_control,
                    'tipo_de_ativo': item.tipo_de_ativo,
                    'funcao_de_seguranca': item.funcao_de_seguranca,
                    'titulo': item.titulo,
                    'descricao': item.descricao,
                    'nist_csf': item.nist_csf,
                    'ig': item.ig,
                    'nome_da_subcategoria': item.nome_da_subcategoria,
                    'acao': temp_actions.get(item.titulo, ''),  # Adiciona a ação correspondente
                })
            messages.success(request, 'Tabela carregada com sucesso!')

            # Log para o carregamento da tabela temporária
            Log.objects.create(
                cliente=request.user,
                descricao=f'O cliente {request.user.first_name} {request.user.last_name} carregou sua tabela.'
            )

            return render(request, 'index.html', {
                'data': data_with_actions,  # Passa os dados com ações
                'column_names': column_names,
            })
        
        except Cliente.DoesNotExist:
            messages.error(request, 'Cliente não encontrado.')
            return redirect('login')
    else:
        return redirect('login')